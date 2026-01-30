# app/services/importer.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import re
import unicodedata

from openpyxl import load_workbook

class ImportErrorExcel(ValueError):
    """Erro de importação/validação do Excel (mensagem amigável)."""


@dataclass
class ImportResult:
    requests: List[Dict[str, Any]]
    source_file: str
    total_rows_read: int
    total_rows_used: int
    warnings: List[str]


# =========================================================
# ✅ PADRÃO OFICIAL (os cabeçalhos do seu Excel)
# (não vão mudar)
# =========================================================
# Empresa | Exercício | Trimestre | Campo/Bloco | Fase | Status | Versão |
# Seção.Expurgo | Def.Projeto | Data Início | Bidround proposto | RIT

# Mapeamento: header "humano" -> chave interna
HEADER_MAP_OFFICIAL = {
    "empresa": "empresa",
    "exercicio": "exercicio",
    "trimestre": "trimestre",
    "campobloco": "campo",
    "fase": "fase",
    "status": "status",
    "versao": "versao",
    "secaoexpurgo": "secao",
    "defprojeto": "defprojeto",
    "datainicio": "datainicio",
    "bidroundproposto": "bidround",
    "rit": "rit",
}

REQUIRED_INTERNAL = [
    "empresa",
    "exercicio",
    "trimestre",
    "campo",
    "fase",
    "status",
    "versao",
    "secao",
    "defprojeto",
    "datainicio",
    "bidround",
    "rit",
]

_DATE_DDMMAAAA = re.compile(r"^\d{8}$")


# =========================
# Normalizações
# =========================
def _norm_header(s: Any) -> str:
    """
    Normaliza cabeçalho para comparação (case-insensitive, sem acentos e sem separadores).
    Exemplos:
      "Exercício" -> "exercicio"
      "Campo/Bloco" -> "campobloco"
      "Seção.Expurgo" -> "secaoexpurgo"
      "Def.Projeto" -> "defprojeto"
      "Data Início" -> "datainicio"
      "Bidround proposto" -> "bidroundproposto"
      "RIT" -> "rit"
    """
    raw = str(s or "").strip().lower()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))

    # remove separadores
    for ch in [" ", ".", "/", "-", "_"]:
        raw = raw.replace(ch, "")
    return raw


def _cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    # Excel às vezes vem float para inteiros
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# =========================
# Regras específicas
# =========================
def _parse_rit(v: Any) -> bool:
    """
    Regra do projeto:
      - vazio -> False
      - "X" (case-insensitive) -> True
    """
    s = _cell_to_str(v).strip().upper()
    if not s:
        return False
    if s == "X":
        return True
    raise ImportErrorExcel("Coluna 'RIT' fora do padrão: use vazio (não) ou 'X' (sim).")


def _excel_date_to_ddmmaaaa(v: Any) -> str:
    """
    datainicio final deve ser ddmmaaaa.
    Aceita:
      - datetime/date (Excel)
      - string ddmmaaaa
      - string "dd/mm/aaaa" ou "dd-mm-aaaa" ou "dd.mm.aaaa"
    """
    if v is None:
        return ""

    if isinstance(v, datetime):
        return v.strftime("%d%m%Y")
    if isinstance(v, date):
        return v.strftime("%d%m%Y")

    s = _cell_to_str(v)
    if not s:
        return ""

    if _DATE_DDMMAAAA.match(s):
        return s

    # tolerância: dd/mm/aaaa, dd-mm-aaaa, dd.mm.aaaa
    s2 = s.replace("/", "").replace("-", "").replace(".", "")
    if _DATE_DDMMAAAA.match(s2):
        return s2

    raise ImportErrorExcel("Coluna 'Data Início' fora do padrão: use data Excel ou ddmmaaaa.")


def _is_row_empty(row_obj: Dict[str, Any]) -> bool:
    """
    Linha “vazia” = todos campos vazios e rit False.
    """
    if bool(row_obj.get("rit", False)):
        return False
    for k, v in row_obj.items():
        if k == "rit":
            continue
        if str(v or "").strip() != "":
            return False
    return True


def _build_header_index_map(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    """
    Constrói o mapa: chave interna -> índice da coluna na planilha,
    baseado no padrão oficial.
    """
    norm_to_idx: Dict[str, int] = {}
    for i, h in enumerate(header_row):
        nh = _norm_header(h)
        if nh:
            norm_to_idx[nh] = i

    internal_to_idx: Dict[str, int] = {}
    for norm_name, internal_name in HEADER_MAP_OFFICIAL.items():
        if norm_name in norm_to_idx:
            internal_to_idx[internal_name] = norm_to_idx[norm_name]

    missing = [k for k in REQUIRED_INTERNAL if k not in internal_to_idx]
    if missing:
        raise ImportErrorExcel(
            "Planilha fora do padrão. Colunas obrigatórias ausentes: " + ", ".join(missing)
        )

    return internal_to_idx


# =========================
# API principal
# =========================
def import_requests_from_excel(xlsx_path: str | Path, sheet_name: Optional[str] = None) -> ImportResult:
    """
    Lê um .xlsx/.xlsm e retorna lista de requests para o requests.json.

    ✅ Padrão oficial do Excel (cabeçalhos humanos) fixo.
    ✅ Sem limite de linhas.
    ✅ Ignora linhas totalmente vazias.
    """
    p = Path(xlsx_path).expanduser().resolve()
    if not p.exists():
        raise ImportErrorExcel(f"Arquivo não encontrado: {p}")
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ImportErrorExcel("Arquivo inválido. Selecione uma planilha Excel (.xlsx/.xlsm).")

    wb = load_workbook(filename=str(p), data_only=True, read_only=True)

    # decide aba
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ImportErrorExcel(f"Aba '{sheet_name}' não encontrada. Abas disponíveis: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ImportErrorExcel("Planilha vazia (sem cabeçalho).")

    idx_map = _build_header_index_map(header_row)

    def get(row: Tuple[Any, ...], key: str) -> Any:
        idx = idx_map[key]
        return row[idx] if idx < len(row) else None

    warnings: List[str] = []
    requests: List[Dict[str, Any]] = []

    total_read = 0
    for ridx, row in enumerate(rows_iter, start=2):
        total_read += 1

        try:
            row_obj = {
                "empresa": _cell_to_str(get(row, "empresa")),
                "exercicio": _cell_to_str(get(row, "exercicio")),
                "trimestre": _cell_to_str(get(row, "trimestre")),
                "campo": _cell_to_str(get(row, "campo")),
                "fase": _cell_to_str(get(row, "fase")),
                "status": _cell_to_str(get(row, "status")),
                "versao": _cell_to_str(get(row, "versao")),
                "secao": _cell_to_str(get(row, "secao")),
                "defprojeto": _cell_to_str(get(row, "defprojeto")),
                "datainicio": _excel_date_to_ddmmaaaa(get(row, "datainicio")),
                "bidround": _cell_to_str(get(row, "bidround")),
                "rit": _parse_rit(get(row, "rit")),
            }
        except ImportErrorExcel as e:
            raise ImportErrorExcel(f"Linha {ridx}: {e}") from None

        if _is_row_empty(row_obj):
            continue

        requests.append(row_obj)

    if not requests:
        raise ImportErrorExcel("Nenhuma linha preenchida encontrada na planilha (todas vazias).")

    return ImportResult(
        requests=requests,
        source_file=str(p),
        total_rows_read=total_read,
        total_rows_used=len(requests),
        warnings=warnings,
    )
